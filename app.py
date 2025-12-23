from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import Workbook, load_workbook
from pathlib import Path
from datetime import datetime

from schemas import schemas, save_schemas

app = FastAPI(title="Excel helper API.")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "*"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# Excel fayllar saqlanadigan papka
BASE_DIR = Path("files")
BASE_DIR.mkdir(exist_ok=True)

@app.get("/")
def root():
    return {"message": "Server is live!"}

@app.get("/files")
def list_files():
    return [f.name for f in BASE_DIR.iterdir() if f.suffix == ".xlsx"]

@app.get("/file/download/{filename}")
def download_excel(filename: str):
    file_path = BASE_DIR / filename

    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")

    return FileResponse(
        path=file_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.post("/create")
def create_file(data: dict):
    filename = f"{data['document_name']}.xlsx"
    file_path = BASE_DIR / filename

    if file_path.exists():
        raise HTTPException(status_code=400, detail="File already exists")

    wb = Workbook()
    wb.remove(wb.active)

    for sheet in data["sheets"]:
        ws = wb.create_sheet(title=sheet["name"])

        # 🔹 HEADER
        headers = [col["name"] for col in sheet["columns"]]
        ws.append(headers)

        # 🔹 TYPE MAPPING (2-qatordan boshlab)
        for col_index, col in enumerate(sheet["columns"], start=1):
            col_letter = ws.cell(row=1, column=col_index).column_letter

            if col["type"] == "number":
                ws.column_dimensions[col_letter].width = 15

            elif col["type"] == "date":
                ws.column_dimensions[col_letter].width = 18

            else:  # text
                ws.column_dimensions[col_letter].width = 20

    wb.save(file_path)

    schemas[filename] = data
    save_schemas(schemas)

    return {
        "message": "Created successfully",
        "file": filename
    }

@app.delete("/delete/{filename}")
def delete_file(filename: str):
    file_path = BASE_DIR / filename

    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")

    # Excel faylni o‘chiramiz
    file_path.unlink()

    # Schema’ni ham o‘chiramiz
    schemas.pop(filename, None)
    save_schemas(schemas)

    return {"message": "Deleted successfully"}

@app.get("/schema/{filename}")
def get_schema(filename: str):
    if filename not in schemas:
        raise HTTPException(status_code=404, detail="Schema not found")

    return schemas[filename]

@app.post("/submit")
def submit(data: dict):
    """
    data = {
        "doc_name": "Test",
        "sheet": "Sheet1",
        "datas": ["Text", "23", "2025-12-10"]
    }
    """

    file_path = Path("files") / f"{data['doc_name']}.xlsx"

    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Excel file not found")

    # schema ni olamiz
    if f"{data['doc_name']}.xlsx" not in schemas:
        raise HTTPException(status_code=404, detail="Schema not found")

    schema = schemas[f"{data['doc_name']}.xlsx"]

    # sheet schema
    sheet_schema = next(
        (s for s in schema["sheets"] if s["name"] == data["sheet"]),
        None
    )

    if not sheet_schema:
        raise HTTPException(status_code=404, detail="Sheet not found in schema")

    wb = load_workbook(file_path)
    ws = wb[data["sheet"]]

    print("📄 SHEET:", ws.title)
    print("📊 BEFORE ROWS:", ws.max_row)

    # 🔁 TYPE GA QARAB TO‘G‘RI CONVERT
    row = []
    for value, col_schema in zip(data["datas"], sheet_schema["columns"]):
        col_type = col_schema["type"]

        if col_type == "number":
            try:
                value = int(value)
            except (ValueError, TypeError):
                value = None

        elif col_type == "date":
            try:
                value = datetime.strptime(value, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                value = None

        # text → o‘z holicha qoladi
        row.append(value)

    # qaysi qatorga yozilishini oldindan bilib olamiz
    next_row = ws.max_row + 1

    ws.append(row)

    # 🔧 FORMATLARNI FAQAT SHU QATORGA BERAMIZ
    for col_index, col_schema in enumerate(sheet_schema["columns"], start=1):
        if col_schema["type"] == "date":
            ws.cell(row=next_row, column=col_index).number_format = "YYYY-MM-DD"

    print("📊 AFTER ROWS:", ws.max_row)

    wb.save(file_path)
    wb.close()

    return {"message": "Success"}